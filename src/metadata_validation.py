"""
Metadata validation module for Excel-based sequencing metadata files.
This module provides validation, detection, and highlighting of issues in metadata.
"""
import os
import re
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except Exception:
    openpyxl = None


def validate_metadata_and_write_report(metadata_file, out_xlsx=None):
    """Validate metadata Excel and write a highlighted copy + change list.

    - out_xlsx: path to write validation workbook (xlsx). If openpyxl not available
      a plain text log will be written to logs/metadata_validation.txt instead.
    """
    issues = []
    seq_pattern = re.compile(r'^[ACGTNacgtn]+$')

    def _is_sequence_series(series):
        vals = [str(v).strip() for v in series if not pd.isna(v)]
        vals = [v for v in vals if v and v.lower() != 'nan' and v.lower() != 'none']
        if not vals:
            return False
        return all(seq_pattern.match(v) for v in vals)

    def _norm_key(val):
        if pd.isna(val):
            return None
        try:
            num = float(val)
            if num.is_integer():
                return str(int(num))
        except Exception:
            pass
        return str(val).strip()

    def _norm_project_name(val):
        """Normalize a project name: replace spaces and underscores with a single underscore."""
        if pd.isna(val):
            return ''
        s = str(val).strip()
        if not s or s.lower() in ('nan', 'none'):
            return ''
        return re.sub(r'[\s_]+', '_', s).strip('_')

    if not metadata_file or not os.path.exists(metadata_file):
        issues.append({'sheet': '', 'row': '', 'col': '', 'message': f'Metadata file not found: {metadata_file}'})
        os.makedirs('metadata', exist_ok=True)
        with open('metadata/metadata_validation.txt', 'w') as tf:
            for it in issues:
                tf.write(f"{it['sheet']}: {it['row']} {it['col']} - {it['message']}\n")
        return

    try:
        xlf = pd.ExcelFile(metadata_file)
    except Exception as e:
        issues.append({'sheet': '', 'row': '', 'col': '', 'message': f'Could not open Excel: {e}'})
        os.makedirs('metadata', exist_ok=True)
        with open('metadata/metadata_validation.txt', 'w') as tf:
            for it in issues:
                tf.write(f"{it['sheet']}: {it['row']} {it['col']} - {it['message']}\n")
        return

    sheet_dfs = {}
    header_rows = {}
    barcode_len_by_lane_group = {}

    for sheet in xlf.sheet_names:
        try:
            raw = pd.read_excel(metadata_file, sheet_name=sheet, header=None)
        except Exception:
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': 'Could not read sheet'})
            continue

        # Find header row BEFORE filling (to avoid filling header rows into data rows)
        header_row = -1
        nrows = raw.shape[0]
        header_keywords = ['Lane', 'Sample_Project', 'Project name', 'Sample Name', 'Sample_Name', 'Sample_ID', 'Lab ID', 'Order ID', 'order ID', 'Email', 'Group', 'group', 'Gr', 'gr']
        for i in range(nrows):
            row = raw.iloc[i]
            row_vals = [str(x) if not pd.isna(x) else '' for x in row.values]

            # Direct match in single row
            if 'Lane' in row_vals and any(k in row_vals for k in ['Sample_Project', 'Project name', 'Sample Name', 'Sample_Name']):
                header_row = i
                break

            # Try combining with the next one or two rows to handle split/multi-line headers
            combined_vals = list(row_vals)
            for j in (1, 2):
                if i + j < nrows:
                    next_row = raw.iloc[i + j]
                    next_vals = [str(x) if not pd.isna(x) else '' for x in next_row.values]
                    combined_vals.extend(next_vals)
                    if 'Lane' in combined_vals and any(k in combined_vals for k in ['Sample_Project', 'Project name', 'Sample Name', 'Sample_Name', 'Lab ID', 'Order ID', 'Email', 'group', 'Group']):
                        # Prefer the lower row index if it contains the majority of header keywords
                        matches_curr = sum(1 for k in header_keywords if k in row_vals)
                        matches_next = sum(1 for k in header_keywords if k in next_vals)
                        header_row = i + j if matches_next >= matches_curr else i
                        break
                else:
                    break
            if header_row != -1:
                break

        if header_row == -1:
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': 'Header row not found'})
            # store raw as-is to copy into workbook
            df_sheet = raw.copy()
            sheet_dfs[sheet] = df_sheet
            header_rows[sheet] = None
            continue

        try:
            df = pd.read_excel(metadata_file, sheet_name=sheet, header=header_row)
        except Exception as e:
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': f'Error parsing sheet with detected header: {e}'})
            sheet_dfs[sheet] = raw
            header_rows[sheet] = None
            continue

        # After reading with proper header, fill down missing values in key columns
        # This handles merged cells in the Excel file where multiple rows share the same project/sample info
        try:
            # Identify columns that should be filled down (metadata columns, not barcode sequences)
            fill_cols = ['Lane', 'Lane.1', 'Group', 'group', 'Gr', 'gr', 'Order ID', 'order ID', 'LabID', 'Lab ID', 'Contact', 'Email',
                        'Project name', 'Project', 'Sample_Project', 'Sample Name', 'Sample_Name',
                        'Sample_ID', 'Index Name']
            cols_to_fill = [c for c in fill_cols if c in df.columns]

            # Only forward-fill i7/index (first index) barcode columns when they contain sequence values.
            # Do NOT forward-fill i5/index2 (second index) columns to avoid propagating blanks for single-index libraries.
            seq_candidate_cols = ['index', 'i7 Barcode Sequence', 'Index']
            seq_cols = [c for c in seq_candidate_cols if c in df.columns and _is_sequence_series(df[c])]

            # Barcode List often leaves Parse/10x groups blank on purpose; do not
            # forward-fill sequence columns there across groups.
            is_barcode_list_sheet = False
            try:
                is_barcode_list_sheet = isinstance(sheet, str) and sheet.strip().lower().startswith('barcode')
            except Exception:
                is_barcode_list_sheet = False
            if is_barcode_list_sheet:
                seq_cols = []

            # Determine if this sheet is a Summary-like sheet; summary sheets should not
            # have Order ID forward-filled — instead, prefer per-row Lab ID values.
            is_summary_sheet = False
            try:
                if isinstance(sheet, str) and 'summary' in sheet.lower():
                    is_summary_sheet = True
            except Exception:
                is_summary_sheet = False

            # If summary sheet, exclude Order ID from forward-fill to avoid copying
            # one Order ID down the whole column; otherwise include it.
            if is_summary_sheet:
                cols_to_fill = [c for c in cols_to_fill if c not in ('Order ID', 'order ID')]

            final_fill = cols_to_fill + seq_cols
            if final_fill:
                df[final_fill] = df[final_fill].ffill()

            # After forward-fill: handle Order ID fills.
            try:
                lab_cols = []
                if 'Lab ID' in df.columns:
                    lab_cols.append('Lab ID')
                if 'LabID' in df.columns:
                    lab_cols.append('LabID')

                _oid_col = next((c for c in ('Order ID', 'order ID') if c in df.columns), None)
                if _oid_col and lab_cols:
                    if is_summary_sheet:
                        # For summary sheets, fill missing Order ID per-row from Lab ID values
                        for lc in lab_cols:
                            try:
                                # rows where Order ID is blank but lab id present
                                missing_order = df[_oid_col].isna() | (df[_oid_col].astype(str).str.strip() == '')
                                has_lab = ~(df[lc].isna()) & (df[lc].astype(str).str.strip() != '')
                                to_fill = missing_order & has_lab
                                if to_fill.any():
                                    df.loc[to_fill, _oid_col] = df.loc[to_fill, lc].astype(str).str.strip()
                                # do NOT break: allow other lab cols to supplement remaining blanks
                            except Exception:
                                continue
                    else:
                        # For non-summary sheets, backfill Order ID from Lab ID if entirely missing
                        for lc in lab_cols:
                            try:
                                missing_order = df[_oid_col].isna() | (df[_oid_col].astype(str).str.strip() == '')
                                has_lab = ~(df[lc].isna()) & (df[lc].astype(str).str.strip() != '')
                                to_fill = missing_order & has_lab
                                if to_fill.any():
                                    df.loc[to_fill, _oid_col] = df.loc[to_fill, lc].astype(str).str.strip()
                                    break
                            except Exception:
                                continue
            except Exception:
                pass
        except Exception:
            # If filling fails, continue with unfilled data
            pass

        # Make Sample_Name values unique within each project by appending suffixes
        # This handles cases where multiple barcodes belong to the same sample (merged cells)
        try:
            project_col = None
            if 'Project' in df.columns:
                project_col = 'Project'
            elif 'Project name' in df.columns:
                project_col = 'Project name'
            elif 'Sample_Project' in df.columns:
                project_col = 'Sample_Project'
            
            sample_name_col = None
            if 'Sample_Name' in df.columns:
                sample_name_col = 'Sample_Name'
            elif 'Sample Name' in df.columns:
                sample_name_col = 'Sample Name'
            
            # If we have both project and sample name columns, make sample names unique
            if project_col and sample_name_col:
                for project in df[project_col].unique():
                    if pd.isna(project) or str(project).strip() == '' or str(project).lower() == 'nan':
                        continue
                    
                    project_mask = df[project_col] == project
                    project_indices = df[project_mask].index
                    
                    # Count occurrences of each Sample_Name within this project
                    sample_names_in_project = df.loc[project_indices, sample_name_col]
                    sample_name_counts = sample_names_in_project.value_counts()
                    
                    # For Sample_Names that appear more than once, add suffixes
                    for sample_name, count in sample_name_counts.items():
                        if count > 1 and pd.notna(sample_name):
                            # Find all occurrences of this Sample_Name in this project
                            dup_mask = (df[project_col] == project) & (df[sample_name_col] == sample_name)
                            dup_indices = df[dup_mask].index
                            
                            # Append suffix to each duplicate (_1, _2, etc.)
                            for i, idx in enumerate(dup_indices, start=1):
                                df.loc[idx, sample_name_col] = f"{sample_name}_{i}"
        except Exception:
            # If uniqueness logic fails, continue with filled but non-unique sample names
            pass

        # Normalize project-name columns (handle minor differences like underscores, extra spaces, case)
        try:
            proj_variants = ['Project name', 'Project Name', 'Project', 'Sample_Project']
            present = [c for c in proj_variants if c in df.columns]
            if present:
                _norm_proj_val = _norm_project_name

                # create normalized versions and detect per-row inconsistencies
                norm_cols = {}
                for c in present:
                    norm_name = f"{c}__norm"
                    df[norm_name] = df[c].apply(_norm_proj_val)
                    norm_cols[c] = norm_name

                for ridx, row in df.iterrows():
                    vals = [row[nc] for nc in norm_cols.values() if row[nc] and str(row[nc]).strip() != '']
                    vals = list(dict.fromkeys(vals))
                    if len(vals) > 1:
                        issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'Project', 'message': 'Inconsistent Project name across columns', 'lane': row.get('Lane', ''), 'group': row.get('Group', row.get('Gr', row.get('group', ''))), 'excel_row': int(ridx) + 2})

                # Create/overwrite a canonical 'Project' column with the first non-empty normalized value
                def _first_nonempty_norm(row):
                    for nc in norm_cols.values():
                        v = row.get(nc)
                        if v and str(v).strip() != '':
                            return v
                    return ''

                df['Project'] = df.apply(_first_nonempty_norm, axis=1)

                # drop temporary norm columns
                try:
                    df.drop(columns=list(norm_cols.values()), inplace=True)
                except Exception:
                    pass
        except Exception:
            pass

        # Build barcode length map by lane/group if barcode sequences are available
        try:
            lane_col = 'Lane' if 'Lane' in df.columns else None
            group_col = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in df.columns), None)

            i7_col = None
            i7_authoritative = False
            if 'i7 Barcode Sequence' in df.columns:
                i7_col = 'i7 Barcode Sequence'
                i7_authoritative = True
            elif 'index' in df.columns:
                i7_col = 'index'
            elif 'Index' in df.columns and _is_sequence_series(df['Index']):
                i7_col = 'Index'

            i5_col = None
            i5_authoritative = False
            if 'i5 Barcode Sequence' in df.columns:
                i5_col = 'i5 Barcode Sequence'
                i5_authoritative = True
            elif 'index2' in df.columns:
                i5_col = 'index2'
            elif 'Index2' in df.columns and _is_sequence_series(df['Index2']):
                i5_col = 'Index2'

            # A sheet is authoritative if it uses dedicated named barcode columns.
            # Authoritative sheets always overwrite; non-authoritative sheets only
            # fill in entries that haven't been set yet, preventing later generic
            # sheets from clobbering lengths established by the Barcode List.
            is_authoritative = i7_authoritative or i5_authoritative

            def _seq_len(series):
                vals = [str(v).strip() for v in series if not pd.isna(v)]
                vals = [v for v in vals if v and v.lower() != 'nan' and v.lower() != 'none']
                if not vals:
                    return 0
                lengths = {len(v) for v in vals}
                return max(lengths)

            if lane_col and group_col and (i7_col or i5_col):
                for (lane_val, group_val), sub in df.groupby([lane_col, group_col]):
                    lane_key = _norm_key(lane_val)
                    group_key = _norm_key(group_val)
                    if lane_key is None or group_key is None:
                        continue
                    i7_len = _seq_len(sub[i7_col]) if i7_col else 0
                    i5_len = _seq_len(sub[i5_col]) if i5_col else 0
                    key = (lane_key, group_key)
                    existing = barcode_len_by_lane_group.get(key)

                    # Keep dedicated barcode sheets as highest priority, but let
                    # other tabs fill missing lengths (common for parse projects).
                    if existing is None:
                        barcode_len_by_lane_group[key] = {
                            'i7_len': i7_len,
                            'i5_len': i5_len
                        }
                    elif is_authoritative:
                        barcode_len_by_lane_group[key] = {
                            'i7_len': i7_len,
                            'i5_len': i5_len
                        }
                    else:
                        merged_i7 = existing.get('i7_len', 0)
                        merged_i5 = existing.get('i5_len', 0)
                        if merged_i7 == 0 and i7_len > 0:
                            merged_i7 = i7_len
                        if merged_i5 == 0 and i5_len > 0:
                            merged_i5 = i5_len
                        barcode_len_by_lane_group[key] = {
                            'i7_len': merged_i7,
                            'i5_len': merged_i5
                        }
        except Exception:
            pass

        sheet_dfs[sheet] = df
        header_rows[sheet] = header_row

        # Basic checks
        if 'Lane' not in df.columns:
            issues.append({'sheet': sheet, 'row': '', 'col': 'Lane', 'message': 'Missing Lane column'})
        else:
            # non-numeric lanes
            try:
                bad_lane = df[pd.to_numeric(df['Lane'], errors='coerce').isna()]
                if not bad_lane.empty:
                    for ridx in bad_lane.index.tolist():
                        issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'Lane', 'message': 'Non-numeric Lane value'})
            except Exception:
                pass

        # Project column existence
        if not any(c in df.columns for c in ['Sample_Project', 'Project name', 'Project']):
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': 'Missing Project column (Sample_Project or Project name)'})

        # Duplicate combined barcodes (only flag if on same lane)
        if 'index' in df.columns and 'Lane' in df.columns:
            idx1 = df['index'].fillna('').astype(str)
            idx2 = df['index2'].fillna('').astype(str) if 'index2' in df.columns else pd.Series([''] * len(df))
            combined = (idx1 + ':' + idx2).replace('nan', '')
            lanes = df['Lane'].fillna('').astype(str)
            
            # Find duplicate barcodes on same lane
            for lane in lanes.unique():
                if lane == '' or pd.isna(lane):
                    continue
                lane_mask = (lanes == lane)
                lane_combined = combined[lane_mask]
                dup_in_lane = lane_combined.duplicated(keep=False) & (lane_combined != ':')
                if dup_in_lane.any():
                    lane_dup_idxs = df.index[lane_mask][dup_in_lane].tolist()
                    for ridx in lane_dup_idxs:
                        issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'index', 'message': 'Duplicate barcode combination (index+index2) on same lane'})

        # Missing indexes when others exist
        if 'index' in df.columns:
            has_any = df['index'].notna() & (df['index'].astype(str).str.strip() != '')
            if has_any.any() and (~has_any).any():
                for ridx in df.index[~has_any].tolist():
                    issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'index', 'message': 'Missing index while other rows have index'})

        # Masking validation runs after all sheets are processed

    # Validate project-name parity between Summary and non-Summary tabs
    _clean_project_name = _norm_project_name

    summary_projects = set()
    non_summary_projects = set()
    non_summary_project_sheets = {}
    summary_proj_orig = {}
    non_summary_proj_orig = {}

    for sheet, df in sheet_dfs.items():
        proj_col = None
        for c in ('Project', 'Project Name', 'Project name', 'Sample_Project'):
            if c in df.columns:
                proj_col = c
                break
        if proj_col is None:
            continue

        is_summary_tab = False
        is_barcode_list_tab = False
        try:
            is_summary_tab = isinstance(sheet, str) and 'summary' in sheet.lower()
            is_barcode_list_tab = isinstance(sheet, str) and 'barcode' in sheet.lower() and 'list' in sheet.lower()
        except Exception:
            pass

        # Only Summary and Barcode List tabs canonically carry project names;
        # all other specialty tabs (10x, BD, etc.) link to Summary via lane/group/Lab ID.
        if not is_summary_tab and not is_barcode_list_tab:
            continue

        for raw_proj in df[proj_col].tolist():
            proj = _clean_project_name(raw_proj)
            if not proj:
                continue
            raw_str = str(raw_proj).strip() if not pd.isna(raw_proj) else ''
            if is_summary_tab:
                summary_projects.add(proj)
                if proj not in summary_proj_orig:
                    summary_proj_orig[proj] = raw_str
            else:
                non_summary_projects.add(proj)
                non_summary_project_sheets.setdefault(proj, set()).add(sheet)
                if proj not in non_summary_proj_orig:
                    non_summary_proj_orig[proj] = raw_str

    # Only compare if Barcode List (the canonical non-summary project-name source) has data.
    if non_summary_projects:
        missing_in_tabs = sorted(summary_projects - non_summary_projects)
        for proj in missing_in_tabs:
            orig = summary_proj_orig.get(proj, proj).replace(' ', '_')
            issues.append({
                'sheet': 'Summary',
                'row': '',
                'col': 'Project',
                'message': f"Project listed in Summary but missing from Barcode List: {orig}"
            })

    missing_in_summary = sorted(non_summary_projects - summary_projects)
    for proj in missing_in_summary:
        orig = non_summary_proj_orig.get(proj, proj).replace(' ', '_')
        sheets = sorted(non_summary_project_sheets.get(proj, set()))
        sheet_list = ', '.join(sheets) if sheets else 'unknown sheet'
        issues.append({
            'sheet': 'Summary',
            'row': '',
            'col': 'Project',
            'message': f"Project present in Barcode List but missing from Summary: {orig}"
        })

    # Validate 'Sample sheet tab' column in Summary against actual lane/group sheet placements.
    # Build two maps:
    #   _lane_group_to_sheets : (lane, group) -> set of sheet names that contain it
    #   _sheet_to_lane_groups : sheet name    -> set of (lane, group) pairs it contains
    # Prefer the right-side duplicate Lane column (renamed "Lane.1" by pandas) because it
    # holds the project-metadata lane matching the Summary, not the BCL sample-sheet lane.
    _lane_group_to_sheets = {}
    _sheet_to_lane_groups = {}
    for _sheet, _df in sheet_dfs.items():
        try:
            if isinstance(_sheet, str) and 'summary' in _sheet.lower():
                continue
            _lc = 'Lane.1' if 'Lane.1' in _df.columns else ('Lane' if 'Lane' in _df.columns else None)
            _gc = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in _df.columns), None)
            if not _lc or not _gc:
                continue
            _sheet_to_lane_groups[_sheet] = set()
            for (_lv, _gv), _ in _df.groupby([_lc, _gc]):
                _lk = _norm_key(_lv)
                _gk = _norm_key(_gv)
                if _lk and _gk:
                    _lane_group_to_sheets.setdefault((_lk, _gk), set()).add(_sheet)
                    _sheet_to_lane_groups[_sheet].add((_lk, _gk))
        except Exception:
            continue

    def _norm_sheet_ref(s):
        return re.sub(r'[\s_]+', '', str(s)).lower()

    for _sheet, _df in sheet_dfs.items():
        try:
            if not (isinstance(_sheet, str) and 'summary' in _sheet.lower()):
                continue
            _tab_col = next((c for c in _df.columns if str(c).strip().lower() == 'sample sheet tab'), None)
            if _tab_col is None:
                continue
            _lc = 'Lane' if 'Lane' in _df.columns else None
            _gc = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in _df.columns), None)
            if not _lc or not _gc:
                continue
            for _pos, (_ridx, _row) in enumerate(_df.iterrows()):
                _tab_val = _row.get(_tab_col)
                if pd.isna(_tab_val) or str(_tab_val).strip() == '':
                    continue
                _tab_str = str(_tab_val).strip()
                # Skip multi-value or annotation-style cells
                if ',' in _tab_str or 'attachment' in _tab_str.lower():
                    continue
                _lk = _norm_key(_row.get(_lc))
                _gk = _norm_key(_row.get(_gc))
                if _lk is None or _gk is None:
                    continue
                # Resolve the named tab to an actual sheet (space/underscore-insensitive)
                _named_sheet = next(
                    (s for s in _sheet_to_lane_groups if _norm_sheet_ref(s) == _norm_sheet_ref(_tab_str)),
                    None
                )
                if _named_sheet is None:
                    continue  # named tab not found — covered by project-parity check
                # Check whether this lane/group is present in the named tab
                if (_lk, _gk) not in _sheet_to_lane_groups[_named_sheet]:
                    _actual = _lane_group_to_sheets.get((_lk, _gk), set())
                    if _actual:
                        _msg = (
                            f"'Sample sheet tab' lists '{_tab_str}' but lane {_lk} group {_gk} "
                            f"data is in: {', '.join(sorted(_actual))}"
                        )
                    else:
                        _tab_coverage = ', '.join(
                            f"lane {l} group {g}"
                            for l, g in sorted(_sheet_to_lane_groups[_named_sheet])
                        ) or 'no lane/group data'
                        _msg = (
                            f"'Sample sheet tab' lists '{_tab_str}' but lane {_lk} group {_gk} "
                            f"is not in that tab (tab covers: {_tab_coverage})"
                        )
                    issues.append({
                        'sheet': _sheet,
                        'row': int(_ridx),
                        'col': str(_tab_col),
                        'message': _msg,
                        'lane': _lk,
                        'group': _gk,
                        'excel_row': int(_ridx) + 2
                    })
        except Exception:
            continue

    # Validate masking against index lengths after all sheets are processed
    def _mask_len_map(masking_str):
        parts = re.split(r'[;,]', str(masking_str))
        mask_map = {}
        for part in parts:
            part = part.strip()
            if not part:
                continue
            m = re.match(r'^([A-Za-z0-9]+)\s*:\s*(\d+)$', part)
            if m:
                mask_map[m.group(1).upper()] = int(m.group(2))
        return mask_map

    def _get_barcode_len(lane_val, group_val):
        # Try multiple variants of lane/group keys to find a mapping in barcode_len_by_lane_group
        if lane_val is None or group_val is None:
            return None
        lk = _norm_key(lane_val)
        gk = _norm_key(group_val)
        candidates = []
        if lk is not None and gk is not None:
            candidates.append((lk, gk))
        try:
            candidates.append((str(lane_val).strip(), str(group_val).strip()))
        except Exception:
            pass
        try:
            candidates.append((str(float(lane_val)), str(float(group_val))))
        except Exception:
            pass
        for ck in candidates:
            if ck in barcode_len_by_lane_group:
                return barcode_len_by_lane_group.get(ck)
        return None

    def _is_10x_multiome_atac_row(row):
        """Return True when a row appears to be 10xMultiomeATACseq metadata."""
        try:
            name_fields = [
                row.get('Sample sheet tab'),
                row.get('Project'),
                row.get('Project name'),
                row.get('Project Name'),
                row.get('Sample_Project'),
            ]
            for val in name_fields:
                if pd.isna(val):
                    continue
                norm = str(val).replace('_', '').replace(' ', '').strip().lower()
                if '10xmultiomeatacseq' in norm:
                    return True
        except Exception:
            return False
        return False

    def _is_flexbar_row(row):
        """Return True when project or sheet metadata indicates a flexbar workflow."""
        try:
            name_fields = [
                row.get('Sample sheet tab'),
                row.get('Project'),
                row.get('Project name'),
                row.get('Project Name'),
                row.get('Sample_Project'),
            ]
            for val in name_fields:
                if pd.isna(val):
                    continue
                if 'flexbar' in str(val).lower():
                    return True
        except Exception:
            return False
        return False

    for sheet, df in sheet_dfs.items():
        if 'Masking' not in df.columns:
            continue

        # iterate with positional index to avoid label/loc mismatches
        for pos, (ridx, row) in enumerate(df.iterrows()):
            masking_val = row.get('Masking')
            if pd.isna(masking_val) or str(masking_val).strip() == '':
                continue

            mask_map = _mask_len_map(masking_val)
            if not mask_map:
                issues.append({'sheet': sheet, 'row': int(pos), 'col': 'Masking', 'message': 'Unrecognized Masking format', 'lane': '', 'group': '', 'excel_row': int(pos) + 2})
                continue

            i1_len = mask_map.get('I1')
            i2_len = mask_map.get('I2')

            # Determine whether this Summary/DataFrame provides actual sequence values
            summary_has_seq1 = False
            summary_has_seq2 = False
            try:
                if 'index' in df.columns and _is_sequence_series(df['index']):
                    summary_has_seq1 = True
                if 'i7 Barcode Sequence' in df.columns and _is_sequence_series(df['i7 Barcode Sequence']):
                    summary_has_seq1 = True
                if 'Index' in df.columns and _is_sequence_series(df['Index']):
                    summary_has_seq1 = True

                if 'index2' in df.columns and _is_sequence_series(df['index2']):
                    summary_has_seq2 = True
                if 'i5 Barcode Sequence' in df.columns and _is_sequence_series(df['i5 Barcode Sequence']):
                    summary_has_seq2 = True
                if 'Index2' in df.columns and _is_sequence_series(df['Index2']):
                    summary_has_seq2 = True
            except Exception:
                summary_has_seq1 = summary_has_seq1
                summary_has_seq2 = summary_has_seq2

            # Prefer barcode lengths from Barcode List via lane/group mapping
            lane_val = row.get('Lane') if 'Lane' in df.columns else None
            _gcol = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in df.columns), None)
            group_val = row.get(_gcol) if _gcol else None

            mapped = _get_barcode_len(lane_val, group_val)
            lane_key = None
            group_key = None
            if mapped is not None:
                # prefer normalized keys for reporting
                try:
                    lane_key = _norm_key(lane_val)
                    group_key = _norm_key(group_val)
                except Exception:
                    lane_key = lane_val
                    group_key = group_val
                len1 = mapped.get('i7_len', 0)
                len2 = mapped.get('i5_len', 0)
                # If mapped i5 length is zero but Barcode List sheet actually contains i5 sequences,
                # try to recompute lengths directly from the Barcode List sheet as a fail-safe.
                if len2 == 0:
                    try:
                        # find a sheet that looks like Barcode List
                        bdf = None
                        for sname, sdf in sheet_dfs.items():
                            if sname.lower().strip().startswith('barcode') or 'i7 Barcode Sequence' in sdf.columns or 'i5 Barcode Sequence' in sdf.columns:
                                bdf = sdf
                                break
                        if bdf is not None:
                            b_lane_col = 'Lane' if 'Lane' in bdf.columns else None
                            b_group_col = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in bdf.columns), None)
                            b_i5_col = 'i5 Barcode Sequence' if 'i5 Barcode Sequence' in bdf.columns else ('index2' if 'index2' in bdf.columns else ('Index2' if 'Index2' in bdf.columns else None))
                            if b_lane_col and b_group_col and b_i5_col:
                                try:
                                    lk = _norm_key(lane_val)
                                    gk = _norm_key(group_val)
                                    vals = []
                                    for _, brow in bdf.iterrows():
                                        try:
                                            if _norm_key(brow[b_lane_col]) == lk and _norm_key(brow[b_group_col]) == gk:
                                                v = brow[b_i5_col]
                                                if not pd.isna(v):
                                                    sv = str(v).strip()
                                                    if sv and sv.lower() not in ('nan','none'):
                                                        vals.append(sv)
                                        except Exception:
                                            continue
                                    if vals:
                                        len2 = max(len(v) for v in vals)
                                except Exception:
                                    pass
                    except Exception:
                        pass
            # Debug: record mapping lookup details for troubleshooting
            try:
                with open('logs/masking_lookup_debug.txt', 'a') as dbg:
                    dbg.write(f"sheet={sheet} pos={pos} lane_val={lane_val} group_val={group_val} mapped={mapped}\n")
            except Exception:
                pass

            # If no mapping exists, compute index lengths from the Summary row values
            if mapped is None:
                index1 = ''
                index2 = ''
                if 'index' in df.columns:
                    index1 = '' if pd.isna(row.get('index')) else str(row.get('index')).strip()
                elif 'i7 Barcode Sequence' in df.columns:
                    index1 = '' if pd.isna(row.get('i7 Barcode Sequence')) else str(row.get('i7 Barcode Sequence')).strip()
                elif 'Index' in df.columns and _is_sequence_series(df['Index']):
                    index1 = '' if pd.isna(row.get('Index')) else str(row.get('Index')).strip()

                if 'index2' in df.columns:
                    index2 = '' if pd.isna(row.get('index2')) else str(row.get('index2')).strip()
                elif 'i5 Barcode Sequence' in df.columns:
                    index2 = '' if pd.isna(row.get('i5 Barcode Sequence')) else str(row.get('i5 Barcode Sequence')).strip()
                elif 'Index2' in df.columns and _is_sequence_series(df['Index2']):
                    index2 = '' if pd.isna(row.get('Index2')) else str(row.get('Index2')).strip()

                len1 = 0 if index1 in ('', 'nan', 'None') else len(index1)
                len2 = 0 if index2 in ('', 'nan', 'None') else len(index2)

                # No barcode data available for this lane/group — cannot validate masking.
                if len1 == 0 and len2 == 0:
                    continue

            if i1_len is not None and len1 != i1_len:
                issues.append({
                    'sheet': sheet,
                    'row': int(pos),
                    'col': 'Masking',
                    'message': f"Masking I1:{i1_len} does not match index length {len1}",
                    'lane': lane_key if lane_key is not None else lane_val,
                    'group': group_key if group_key is not None else group_val,
                    'excel_row': int(pos) + 2
                })

            if i2_len is not None:
                _tab_ref = row.get('Sample sheet tab', '')
                _tab_ref_str = '' if pd.isna(_tab_ref) else str(_tab_ref).lower()
                is_atac_tab = (
                    (isinstance(sheet, str) and 'atac' in sheet.lower())
                    or 'atac' in _tab_ref_str
                )
                skip_i2_len_check = (
                    len2 == 0
                    and (
                        _is_10x_multiome_atac_row(row)
                        or _is_flexbar_row(row)
                        or is_atac_tab
                    )
                )

                if skip_i2_len_check:
                    continue

                if i2_len == 0 and len2 != 0:
                    issues.append({
                        'sheet': sheet,
                        'row': int(pos),
                        'col': 'Masking',
                        'message': f"Masking I2:0 but index2 length is {len2}",
                        'lane': lane_key if lane_key is not None else lane_val,
                        'group': group_key if group_key is not None else group_val,
                        'excel_row': int(pos) + 2
                    })
                elif i2_len > 0 and len2 != i2_len:
                    issues.append({
                        'sheet': sheet,
                        'row': int(pos),
                        'col': 'Masking',
                        'message': f"Masking I2:{i2_len} does not match index2 length {len2}",
                        'lane': lane_key if lane_key is not None else lane_val,
                        'group': group_key if group_key is not None else group_val,
                        'excel_row': int(pos) + 2
                    })
                elif mapped is None and i2_len > 0 and 'index2' not in df.columns and 'i5 Barcode Sequence' not in df.columns:
                    issues.append({
                        'sheet': sheet,
                        'row': int(pos),
                        'col': 'Masking',
                        'message': f"Masking I2:{i2_len} but index2 column is missing",
                        'lane': lane_key if lane_key is not None else lane_val,
                        'group': group_key if group_key is not None else group_val,
                        'excel_row': int(pos) + 2
                    })

    # Build (lane_key, group_key) -> {Lane, Group, Project, Order ID} lookup for column propagation.
    # Summary sheets are the authoritative source: their values always overwrite any previously set
    # values from non-Summary sheets. Non-Summary sheets only fill keys not already present.
    _prop_lookup = {}

    def _populate_prop_lookup(items, overwrite=False):
        for _sname, _sdf in items:
            _lc = 'Lane.1' if 'Lane.1' in _sdf.columns else ('Lane' if 'Lane' in _sdf.columns else None)
            _gc = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in _sdf.columns), None)
            _pc = next((c for c in ('Project', 'Project name', 'Project Name', 'Sample_Project') if c in _sdf.columns), None)
            _oc = next((c for c in ('Order ID', 'order ID') if c in _sdf.columns), None)
            if not _lc or not _gc:
                continue
            for _, _row in _sdf.iterrows():
                try:
                    _lk = _norm_key(_row.get(_lc))
                    _gk = _norm_key(_row.get(_gc))
                    if not _lk or not _gk:
                        continue
                    entry = _prop_lookup.setdefault((_lk, _gk), {'Lane': _lk, 'Group': _gk})
                    if _pc:
                        v = _row.get(_pc)
                        if not pd.isna(v) and str(v).strip() not in ('', 'nan', 'None'):
                            if overwrite or 'Project' not in entry:
                                entry['Project'] = str(v).strip()
                    if _oc:
                        v = _row.get(_oc)
                        if not pd.isna(v) and str(v).strip() not in ('', 'nan', 'None'):
                            if overwrite or 'Order ID' not in entry:
                                entry['Order ID'] = str(v).strip()
                except Exception:
                    continue

    # Non-Summary sheets first (fill only); then Summary sheets overwrite with authoritative values.
    _populate_prop_lookup(
        ((s, d) for s, d in sheet_dfs.items()
         if not (isinstance(s, str) and 'summary' in s.lower())),
        overwrite=False,
    )
    _populate_prop_lookup(
        ((s, d) for s, d in sheet_dfs.items()
         if isinstance(s, str) and 'summary' in s.lower()),
        overwrite=True,
    )

    # Propagate Lane, Group, Project, Order ID to every sheet.
    # RECOMMENDED_CHANGES and RC_ORIENTATION are written separately and never appear in sheet_dfs,
    # but guard against them explicitly in case the input file already contains those names.
    _NO_PROPAGATE = {'RECOMMENDED_CHANGES', 'RC_ORIENTATION'}
    for _sheet in list(sheet_dfs.keys()):
        if _sheet in _NO_PROPAGATE:
            continue
        _df = sheet_dfs[_sheet]
        _lc = 'Lane.1' if 'Lane.1' in _df.columns else ('Lane' if 'Lane' in _df.columns else None)
        _gc = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in _df.columns), None)
        if not _lc or not _gc:
            continue
        # Each entry: (lookup_key, [aliases in priority order])
        # The first alias found in the sheet is used; if none found, the first alias is created.
        _PROP_TARGETS = [
            ('Lane',     ['Lane']),
            ('Group',    ['Group', 'Gr', 'group', 'gr']),
            ('Project',  ['Project', 'Project name', 'Project Name', 'Sample_Project']),
            ('Order ID', ['Order ID', 'order ID']),
        ]
        _newly_added = []
        for _info_key, _aliases in _PROP_TARGETS:
            if _info_key == 'Lane' and 'Lane.1' in _df.columns:
                continue

            _target_col = next((c for c in _aliases if c in _df.columns), None)
            _col_absent = _target_col is None
            if _col_absent:
                _target_col = _aliases[0]
                _df[_target_col] = ''
                _missing = pd.Series([True] * len(_df), index=_df.index)
            else:
                _missing = _df[_target_col].isna() | (
                    _df[_target_col].astype(str).str.strip().isin(('', 'nan', 'None'))
                )

            if not _missing.any():
                continue

            def _get_val(row, lc=_lc, gc=_gc, ik=_info_key):
                try:
                    lv = _norm_key(row.get(lc))
                    # Lane and Lane.1 are aliases; fall back to the other if one is absent
                    if lv is None:
                        alt = 'Lane' if lc == 'Lane.1' else ('Lane.1' if lc == 'Lane' else None)
                        if alt:
                            lv = _norm_key(row.get(alt))
                    gv = _norm_key(row.get(gc))
                    if lv and gv:
                        return _prop_lookup.get((lv, gv), {}).get(ik, '')
                except Exception:
                    pass
                return ''

            _df.loc[_missing, _target_col] = _df[_missing].apply(_get_val, axis=1)
            if _col_absent:
                _newly_added.append(_target_col)

        if _newly_added:
            # Move key columns to the front for visibility; keep whatever alias name the sheet uses
            _front = [c for c in _df.columns if c in {'Lane', 'Group', 'Gr', 'group', 'gr',
                                                        'Project', 'Project name', 'Project Name', 'Sample_Project',
                                                        'Order ID', 'order ID'}]
            _rest = [c for c in _df.columns if c not in set(_front)]
            _df = _df[_front + _rest]
        sheet_dfs[_sheet] = _df

    # Write validation workbook if possible
    os.makedirs('metadata', exist_ok=True)
    if out_xlsx is None:
        out_xlsx = os.path.join('metadata', f"metadata_validation_{os.path.basename(metadata_file)}.xlsx")

    if openpyxl is None:
        # fallback: dump issues to text file
        with open('metadata/metadata_validation.txt', 'w') as tf:
            for it in issues:
                tf.write(f"{it['sheet']}: row={it['row']} col={it['col']} - {it['message']}\n")
        return

    # Use pandas to write sheets, then highlight rows with issues
    try:
        with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
            for sheet, df in sheet_dfs.items():
                # If df has numeric header rows (raw), write as-is
                try:
                    df.to_excel(writer, sheet_name=sheet, index=False)
                except Exception:
                    # fallback to writing raw values
                    pd.DataFrame(df).to_excel(writer, sheet_name=sheet, index=False)

            # Write recommended changes sheet
            # Enrich issues with lane/group where possible for clearer RECOMMENDED_CHANGES
            for it in issues:
                sh = it.get('sheet')
                r = it.get('row')
                # Prefer existing values if already provided (e.g., masking checks set these)
                it_lane = it.get('lane', '')
                it_group = it.get('group', '')
                # If missing, try to look up from the sheet by row index
                try:
                    if (not it_lane or pd.isna(it_lane)) and sh in sheet_dfs and isinstance(r, (int, float)):
                        df_lookup = sheet_dfs[sh]
                        if 'Lane' in df_lookup.columns and int(r) in df_lookup.index:
                            it_lane = df_lookup.loc[int(r), 'Lane']
                    if (not it_group or pd.isna(it_group)) and sh in sheet_dfs and isinstance(r, (int, float)):
                        df_lookup = sheet_dfs[sh]
                        for gcol in ('Group', 'Gr', 'group'):
                            if gcol in df_lookup.columns and int(r) in df_lookup.index:
                                it_group = df_lookup.loc[int(r), gcol]
                                break
                except Exception:
                    pass
                it['lane'] = it_lane if it_lane is not None else ''
                it['group'] = it_group if it_group is not None else ''
                # Prefer existing excel_row if present, otherwise compute
                try:
                    if 'excel_row' in it and it.get('excel_row') not in (None, ''):
                        pass
                    else:
                        it['excel_row'] = int(r) + 2 if isinstance(r, (int, float)) else ''
                except Exception:
                    it['excel_row'] = ''

            if issues:
                issues_df = pd.DataFrame(issues)
                try:
                    issues_df['_lane_sort'] = pd.to_numeric(issues_df['lane'], errors='coerce')
                    issues_df['_group_sort'] = issues_df['group'].astype(str)
                    issues_df.sort_values(['_lane_sort', '_group_sort'], inplace=True, na_position='last')
                    issues_df.drop(columns=['_lane_sort', '_group_sort'], inplace=True)
                except Exception:
                    pass
            else:
                issues_df = pd.DataFrame([{'sheet': 'OK', 'row': '', 'col': '', 'message': 'No issues detected'}])
            issues_df.to_excel(writer, sheet_name='RECOMMENDED_CHANGES', index=False)

            # Write RC_ORIENTATION sheet from orientation_decision JSON files
            import json as _json
            import glob as _glob

            # Build project -> group lookup from loaded metadata sheets
            _proj_to_group = {}
            for _sname, _sdf in sheet_dfs.items():
                _proj_col = next((c for c in ('Sample_Project', 'Project', 'project') if c in _sdf.columns), None)
                _grp_col = next((c for c in ('Group', 'Gr', 'group', 'gr') if c in _sdf.columns), None)
                if _proj_col and _grp_col:
                    for _, _row in _sdf.iterrows():
                        try:
                            _p = str(_row[_proj_col]).strip()
                            _g = str(_row[_grp_col]).strip()
                            if _p and _p.lower() not in ('nan', 'none', ''):
                                _proj_to_group[_p] = _g
                        except Exception:
                            pass

            # Supplement lookup from renaming_map CSVs (have Sample_Project + Group columns)
            import csv as _csv
            for _rmap in sorted(_glob.glob('results/renaming_map_*.csv')):
                try:
                    with open(_rmap, newline='') as _f:
                        for _row in _csv.DictReader(_f):
                            _p = _row.get('Sample_Project', '').strip()
                            _g = _row.get('Group', '').strip()
                            if _p and _p.lower() not in ('nan', 'none', ''):
                                _proj_to_group[_p] = _g
                except Exception:
                    pass

            rc_rows = []
            try:
                for dec_file in sorted(_glob.glob('logs/*/orientation_decision_*.json')):
                    config_id = os.path.basename(dec_file).replace('orientation_decision_', '').replace('.json', '')
                    with open(dec_file) as _df:
                        dec = _json.load(_df)
                    for project, orientation in dec.items():
                        group = _proj_to_group.get(project, '')
                        rc_rows.append({'config_id': config_id, 'group': group, 'project': project, 'orientation': orientation})
            except Exception:
                pass
            if rc_rows:
                rc_df = pd.DataFrame(rc_rows, columns=['config_id', 'group', 'project', 'orientation'])
            else:
                rc_df = pd.DataFrame([{'config_id': '', 'group': '', 'project': '', 'orientation': 'No RC decisions found'}])
            rc_df.to_excel(writer, sheet_name='RC_ORIENTATION', index=False)

        # Apply highlighting
        wb = openpyxl.load_workbook(out_xlsx)
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        orange_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        bold = Font(bold=True)

        # Highlight RC rows in RC_ORIENTATION sheet
        if 'RC_ORIENTATION' in wb.sheetnames:
            ws_rc = wb['RC_ORIENTATION']
            for row in ws_rc.iter_rows(min_row=2):
                orient_cell = next((c for c in row if ws_rc.cell(1, c.column).value == 'orientation'), None)
                if orient_cell and orient_cell.value and orient_cell.value.startswith('rc'):
                    for c in row:
                        try:
                            c.fill = orange_fill
                        except Exception:
                            pass

        # Group issues by sheet and highlight entire row where issue occurred
        for it in issues:
            sh = it.get('sheet')
            r = it.get('row')
            if sh in wb.sheetnames and isinstance(r, (int, float)):
                ws = wb[sh]
                # Excel output from pandas has header at row 1 and data starting at row 2.
                # Map DataFrame index -> excel row by adding 2 (1-based excel rows).
                try:
                    excel_row = int(r) + 2
                except Exception:
                    continue
                max_col = ws.max_column
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=excel_row, column=c)
                    try:
                        cell.fill = red_fill
                    except Exception:
                        pass

        # Bold header rows where header detected
        for sh, hr in header_rows.items():
            if sh not in wb.sheetnames:
                continue
            ws = wb[sh]
            # Header written by pandas is at row 1 in the output workbook
            excel_header = 1
            for c in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=excel_header, column=c)
                    cell.font = bold
                except Exception:
                    pass

        wb.save(out_xlsx)
    except Exception as e:
        # final fallback: write issues to text file
        with open('metadata/metadata_validation.txt', 'w') as tf:
            tf.write(f'Error writing validation workbook: {e}\n')
            for it in issues:
                tf.write(f"{it['sheet']}: row={it['row']} col={it['col']} - {it['message']}\n")
    return
